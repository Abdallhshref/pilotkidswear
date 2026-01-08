from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Sum, Count
from orders.models import Order
from store.models import Product

from django.utils import timezone
from datetime import timedelta
from decimal import Decimal

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    orders = Order.objects.all()
    
    # --- Metrics ---
    # 1. Total Orders
    total_orders = orders.count()

    # 2. Approved (Delivered)
    delivered_count = orders.filter(status='delivered').count()

    # 3. Revenue (Lifetime & Monthly)
    # Since we compute total dynamically, we iterate. Optimization: Store 'total_amount' in DB in future.
    revenue_total = Decimal(0)
    confirmed_orders = orders.exclude(status='cancelled')
    for o in confirmed_orders:
        revenue_total += o.get_total_price()

    # Monthly
    now = timezone.now()
    first_day_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    revenue_month = Decimal(0)
    month_orders = confirmed_orders.filter(created_at__gte=first_day_current_month)
    for o in month_orders:
        revenue_month += o.get_total_price()

    # 4. Chart Data (Last 12 months sales)
    # Simple manual aggregation for MVP
    # Format: {'Jan': 100, 'Feb': 200, ...}
    sales_data = [] # List of values for the chart
    months_labels = [] # List of labels
    
    # We want last 12 months including current
    for i in range(11, -1, -1):
        # Calculate start and end of that month
        # Simplified: Just getting the month name and filling 0 for now to keep it fast
        # (Real implementation needs proper date ranges)
        # Check if user wants real data or just the structure. The prompt says "like this with the same data" but "for my website"
        # I will attempt real data extraction
        pass
    
    # Better approach for Chart.js:
    # Just group by month for the current year
    # Or just mock it if no data exists yet?
    # Let's try to get simple data: Sales per month for current year
    
    from django.db.models.functions import TruncMonth
    # Note: Aggregate on dynamic property isn't possible directly with simple ORM unless we have total stored.
    # We'll just pass placeholder data if DB is empty, or simple logic.
    # Since we don't have many orders, let's just make the chart "Sales Dynamics" show valid data.
    
    # Let's iterate last 6 months
    chart_labels = []
    chart_data = []
    
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        next_month_start = (month_start + timedelta(days=32)).replace(day=1)
        month_label = month_start.strftime("%b")
        
        # Filter orders in this range
        m_orders = confirmed_orders.filter(created_at__gte=month_start, created_at__lt=next_month_start)
        m_total = sum(o.get_total_price() for o in m_orders)
        
        chart_labels.append(month_label)
        chart_data.append(float(m_total))

    total_products = Product.objects.count()
    recent_orders = orders.order_by('-created_at')[:10]

    context = {
        'total_orders': total_orders,
        'logo_orders': total_orders, # Using same val for "Orders" card
        'delivered_count': delivered_count,
        'revenue_total': revenue_total,
        'revenue_month': revenue_month,
        'total_products': total_products,
        'recent_orders': recent_orders,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    }
    return render(request, 'core/dashboard.html', context)

# Export functionality
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from collections import defaultdict

@user_passes_test(lambda u: u.is_superuser)
def export_users_excel(request):
    # Aggregate user data from orders
    orders = Order.objects.all()
    user_data = defaultdict(lambda: {'orders': 0, 'total': Decimal(0), 'email': '', 'phone': '', 'city': ''})
    
    for order in orders:
        key = order.full_name
        user_data[key]['orders'] += 1
        user_data[key]['total'] += order.get_total_price()
        user_data[key]['email'] = order.email
        user_data[key]['phone'] = order.phone_number
        user_data[key]['city'] = order.city
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['Name', 'Email', 'Phone', 'City', 'Total Orders', 'Total Spent (LE)']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for name, data in sorted(user_data.items()):
        ws.append([
            name,
            data['email'],
            data['phone'],
            data['city'],
            data['orders'],
            float(data['total'])
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="customers_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@user_passes_test(lambda u: u.is_superuser)
def export_users_csv(request):
    # Aggregate user data
    orders = Order.objects.all()
    user_data = defaultdict(lambda: {'orders': 0, 'total': Decimal(0), 'email': '', 'phone': '', 'city': ''})
    
    for order in orders:
        key = order.full_name
        user_data[key]['orders'] += 1
        user_data[key]['total'] += order.get_total_price()
        user_data[key]['email'] = order.email
        user_data[key]['phone'] = order.phone_number
        user_data[key]['city'] = order.city
    
    # Create CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="customers_export_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Email', 'Phone', 'City', 'Total Orders', 'Total Spent (LE)'])
    
    for name, data in sorted(user_data.items()):
        writer.writerow([
            name,
            data['email'],
            data['phone'],
            data['city'],
            data['orders'],
            float(data['total'])
        ])
    
    return response

@user_passes_test(lambda u: u.is_superuser)
def export_orders_excel(request):
    orders = Order.objects.all().order_by('-created_at')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Headers
    headers = ['Order ID', 'Customer', 'Email', 'Date', 'Items', 'Status', 'Total (LE)', 'Payment Method']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for order in orders:
        # Build items string
        items_list = []
        for item in order.items.all():
            items_list.append(f"{item.variant.product.name} ({item.variant.color.name}/{item.variant.size.name}) x{item.quantity}")
        items_str = "; ".join(items_list)
        
        payment_method = (order.payment_metadata or {}).get('payment_method', 'N/A')
        
        ws.append([
            str(order.tracking_id)[:8],
            order.full_name,
            order.email,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            items_str,
            order.get_status_display(),
            float(order.get_total_price()),
            payment_method.upper()
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@user_passes_test(lambda u: u.is_superuser)
def export_orders_csv(request):
    orders = Order.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Email', 'Date', 'Items', 'Status', 'Total (LE)', 'Payment Method'])
    
    for order in orders:
        items_list = []
        for item in order.items.all():
            items_list.append(f"{item.variant.product.name} ({item.variant.color.name}/{item.variant.size.name}) x{item.quantity}")
        items_str = "; ".join(items_list)
        
        payment_method = (order.payment_metadata or {}).get('payment_method', 'N/A')
        
        writer.writerow([
            str(order.tracking_id)[:8],
            order.full_name,
            order.email,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            items_str,
            order.get_status_display(),
            float(order.get_total_price()),
            payment_method.upper()
        ])
    
    return response
